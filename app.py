import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import datetime
from copy import copy
import plotly.express as px
from io import BytesIO
from pathlib import Path

# --- Cached file reading for performance ---
@st.cache_data
def load_file(file_bytes: bytes, file_name: str, is_stacked: bool = False) -> tuple[pd.DataFrame, int]:
    """Load CSV or Excel file with caching to avoid redundant reads."""
    # Stacked datasets are always stored as CSV
    if is_stacked or file_name.lower().endswith(".csv"):
        return read_flexible_csv(file_bytes)
    else:
        return pd.read_excel(io.BytesIO(file_bytes)), 0

def read_flexible_csv(file_bytes: bytes, max_check_lines: int = 30) -> tuple[pd.DataFrame, int]:
    """
    Read a CSV file that may have extra header or junk lines before the real data.
    Automatically detects which line contains the column headers.
    """
    text = file_bytes.decode(errors="ignore").splitlines()
    
    # If empty file, return empty dataframe
    if not text:
        return pd.DataFrame(), 0
    
    sample_lines = text[:max_check_lines]

    header_row = 0
    for i, line in enumerate(sample_lines):
        parts = [p.strip() for p in line.replace("\t", ",").split(",") if p.strip()]
        if len(parts) < 2:
            continue

        non_numeric = sum(1 for p in parts if not p.replace(".", "").replace("-", "").replace(":", "").replace("/", "").isdigit())
        if non_numeric / len(parts) >= 0.5:
            header_row = i
            break

    try:
        df = pd.read_csv(io.BytesIO(file_bytes), header=header_row, low_memory=False)
        # Remove any completely empty rows at the start
        df = df.dropna(how='all').reset_index(drop=True)
    except Exception:
        df = pd.read_csv(io.BytesIO(file_bytes), header=0, low_memory=False)

    return df, header_row

def detect_datetime_column(df: pd.DataFrame) -> str | None:
    """Find the first column that looks like a datetime column."""
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            return col
        if "date" in col.lower() or "time" in col.lower():
            return col
    return None

def get_selectable_columns(df: pd.DataFrame) -> list[str]:
    """
    Return columns that should be available for selection.
    Excludes datetime columns, index-like columns, and unnamed columns.
    """
    exclude_cols = set()
    
    for col in df.columns:
        col_lower = col.lower().strip()
        
        # Exclude datetime columns
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            exclude_cols.add(col)
            continue
        
        # Exclude columns with date/time in the name
        if "date" in col_lower or "time" in col_lower:
            exclude_cols.add(col)
            continue
        
        # Exclude index-like columns
        if col_lower in ("index", "idx", "id", "row", "row_num", "row_number", "unnamed: 0"):
            exclude_cols.add(col)
            continue
        
        # Exclude any "Unnamed:" columns (pandas auto-generated)
        if col_lower.startswith("unnamed:"):
            exclude_cols.add(col)
            continue
    
    return [c for c in df.columns if c not in exclude_cols]

def is_string_column(series: pd.Series) -> bool:
    """
    Determine if a column should be treated as string data.
    Returns True if the column contains non-numeric string data.
    """
    if series.dtype == object or series.dtype.name == 'string':
        # Try converting to numeric - if most values fail, it's a string column
        numeric_converted = pd.to_numeric(series, errors='coerce')
        non_null_original = series.dropna()
        non_null_numeric = numeric_converted.dropna()
        
        if len(non_null_original) == 0:
            return False
        
        # If less than half converted successfully, treat as string
        conversion_rate = len(non_null_numeric) / len(non_null_original)
        return conversion_rate < 0.5
    
    return False

def check_stack_compatibility(file_data: dict, file_names: list[str]) -> dict:
    """
    Check if files are compatible for stacking.
    Returns dict with 'compatible' bool, 'warnings' list, and 'common_columns' list.
    """
    if len(file_names) < 2:
        return {"compatible": True, "warnings": [], "common_columns": []}
    
    # Load all files and get their columns
    all_columns = []
    all_dfs = []
    
    for fn in file_names:
        df, _ = load_file(file_data[fn]["raw_bytes"], fn)
        selectable = get_selectable_columns(df)
        all_columns.append(set(selectable))
        all_dfs.append((fn, df))
    
    # Find common columns
    common = all_columns[0]
    for cols in all_columns[1:]:
        common = common.intersection(cols)
    
    warnings = []
    
    # Check for column mismatches
    all_cols_union = set()
    for cols in all_columns:
        all_cols_union = all_cols_union.union(cols)
    
    missing_cols = all_cols_union - common
    if missing_cols:
        warnings.append(f"⚠️ Column mismatch: The following columns are not present in all files and will be excluded: {', '.join(sorted(missing_cols))}")
    
    # Check for overlapping timestamps
    time_ranges = []
    for fn, df in all_dfs:
        dt_col = detect_datetime_column(df)
        if dt_col:
            times = pd.to_datetime(df[dt_col], errors='coerce').dropna()
            if len(times) > 0:
                time_ranges.append((fn, times.min(), times.max()))
    
    # Check for overlaps between any pair of files
    overlaps = []
    for i, (fn1, start1, end1) in enumerate(time_ranges):
        for fn2, start2, end2 in time_ranges[i+1:]:
            # Check if ranges overlap
            if start1 <= end2 and start2 <= end1:
                overlap_start = max(start1, start2)
                overlap_end = min(end1, end2)
                overlaps.append(f"{fn1} and {fn2} ({overlap_start.strftime('%Y-%m-%d %H:%M')} to {overlap_end.strftime('%Y-%m-%d %H:%M')})")
    
    if overlaps:
        warnings.append(f"⚠️ Overlapping timestamps detected between: {'; '.join(overlaps)}")
    
    return {
        "compatible": True,  # Still compatible, just with warnings
        "warnings": warnings,
        "common_columns": sorted(list(common)),
        "time_ranges": time_ranges
    }

def create_stacked_dataframe(file_data: dict, file_names: list[str], overlap_handling: str) -> tuple[pd.DataFrame, int]:
    """
    Stack multiple files into a single DataFrame.
    Returns (stacked_df, header_row).
    """
    dfs = []
    
    for fn in file_names:
        df, _ = load_file(file_data[fn]["raw_bytes"], fn)
        
        # Set datetime index
        dt_col = detect_datetime_column(df)
        if dt_col:
            df[dt_col] = pd.to_datetime(df[dt_col], errors='coerce')
            df = df.set_index(dt_col)
        
        # Drop any remaining datetime-like columns or index columns before stacking
        cols_to_keep = []
        for col in df.columns:
            col_lower = col.lower().strip()
            if "date" in col_lower or "time" in col_lower:
                continue
            if col_lower in ("index", "idx", "unnamed: 0") or col_lower.startswith("unnamed:"):
                continue
            cols_to_keep.append(col)
        
        df = df[cols_to_keep]
        dfs.append(df)
    
    # Get common columns only
    common_cols = set(dfs[0].columns)
    for df in dfs[1:]:
        common_cols = common_cols.intersection(set(df.columns))
    common_cols = sorted(list(common_cols))
    
    if not common_cols:
        # If no common columns, return empty dataframe
        return pd.DataFrame(), 0
    
    # Filter to common columns and concatenate
    dfs_filtered = [df[common_cols] for df in dfs]
    stacked = pd.concat(dfs_filtered, axis=0)
    stacked = stacked.sort_index()
    
    # Remove rows with NaT index
    stacked = stacked[stacked.index.notna()]
    
    # Handle overlapping timestamps
    if stacked.index.has_duplicates:
        if overlap_handling == "Keep first occurrence":
            stacked = stacked[~stacked.index.duplicated(keep='first')]
        elif overlap_handling == "Keep last occurrence":
            stacked = stacked[~stacked.index.duplicated(keep='last')]
        elif overlap_handling == "Average values":
            # For averaging, handle numeric and non-numeric separately
            numeric_cols = stacked.select_dtypes(include='number').columns.tolist()
            non_numeric_cols = [c for c in stacked.columns if c not in numeric_cols]
            
            if numeric_cols:
                numeric_avg = stacked[numeric_cols].groupby(stacked.index).mean()
            else:
                numeric_avg = pd.DataFrame(index=stacked.index.unique())
            
            if non_numeric_cols:
                non_numeric_first = stacked[non_numeric_cols].groupby(stacked.index).first()
                stacked = pd.concat([numeric_avg, non_numeric_first], axis=1)
            else:
                stacked = numeric_avg
    
    return stacked, 0

def get_column_types(df: pd.DataFrame, columns: list[str]) -> dict[str, str]:
    """
    Determine the data type for each column.
    Returns a dict mapping column name to 'numeric' or 'string'.
    """
    col_types = {}
    for col in columns:
        if col in df.columns:
            col_types[col] = 'string' if is_string_column(df[col]) else 'numeric'
        else:
            col_types[col] = 'numeric'  # default
    return col_types

def prepare_dataframe(df: pd.DataFrame, dupe_handling: str = "Average values") -> pd.DataFrame:
    """Set datetime index, sort, and handle duplicates."""
    df = df.copy()
    
    dt_col = detect_datetime_column(df)
    if dt_col:
        df[dt_col] = pd.to_datetime(df[dt_col], errors="coerce")
        df = df.set_index(dt_col)
    else:
        df.index = pd.to_datetime(df.index, errors="coerce")
    
    df = df.sort_index()
    df = df[df.index.notna()]
    
    if df.index.has_duplicates:
        # Separate numeric and non-numeric columns
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        non_numeric_cols = [c for c in df.columns if c not in numeric_cols]
        
        if dupe_handling == "Average values":
            if numeric_cols:
                numeric_df = df[numeric_cols].groupby(df.index).mean()
            else:
                numeric_df = pd.DataFrame(index=df.index.unique())
            
            if non_numeric_cols:
                non_numeric_df = df[non_numeric_cols].groupby(df.index).first()
                df = pd.concat([numeric_df, non_numeric_df], axis=1)
            else:
                df = numeric_df
                
        elif dupe_handling == "Maximum value":
            if numeric_cols:
                numeric_df = df[numeric_cols].groupby(df.index).max()
            else:
                numeric_df = pd.DataFrame(index=df.index.unique())
            
            if non_numeric_cols:
                non_numeric_df = df[non_numeric_cols].groupby(df.index).first()
                df = pd.concat([numeric_df, non_numeric_df], axis=1)
            else:
                df = numeric_df
                
        elif dupe_handling == "Minimum value":
            if numeric_cols:
                numeric_df = df[numeric_cols].groupby(df.index).min()
            else:
                numeric_df = pd.DataFrame(index=df.index.unique())
            
            if non_numeric_cols:
                non_numeric_df = df[non_numeric_cols].groupby(df.index).first()
                df = pd.concat([numeric_df, non_numeric_df], axis=1)
            else:
                df = numeric_df
    
    return df

def apply_cleanup(df: pd.DataFrame, col: str, method: str, is_string: bool = False) -> pd.DataFrame:
    """Apply missing data handling to a specific column."""
    df = df.copy()
    
    if is_string:
        # For string columns, only certain cleanup methods make sense
        if method == "Fill using last valid entry":
            df[col] = df[col].ffill()
        elif method == "Fill using next available value":
            df[col] = df[col].bfill()
        elif method == "Delete the entire row of data":
            df = df[df[col].notna() & (df[col] != '')]
        elif method == "Fill with zero":
            df[col] = df[col].fillna('')  # Use empty string for strings
        # Linear interpolation doesn't apply to strings - use ffill instead
        elif method == "Fill with a linear interpolation between the nearest values":
            df[col] = df[col].ffill()
    else:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        
        if method == "Fill using last valid entry":
            df[col] = df[col].ffill()
        elif method == "Fill using next available value":
            df[col] = df[col].bfill()
        elif method == "Fill with a linear interpolation between the nearest values":
            df[col] = df[col].interpolate(method="time")
        elif method == "Delete the entire row of data":
            df = df[df[col].notna()]
        elif method == "Fill with zero":
            df[col] = df[col].fillna(0)
    
    return df

# --- Page Config ---
st.set_page_config(page_title="Merge-o-matic 2000", layout="wide")
st.title("Merge-o-matic 2000 🤖")
st.write("Upload CSV or Excel files with data and timestamps. We'll time-align and filter them and let you download a combined dataset.")

# --- Initialize Session State ---
if "file_data" not in st.session_state:
    st.session_state.file_data = {}
if "graph_fig" not in st.session_state:
    st.session_state.graph_fig = None
if "stacked_datasets" not in st.session_state:
    st.session_state.stacked_datasets = {}  # name -> {files: [], overlap_handling: str}

# --- Step 1: File Upload ---
st.header("📁 Upload CSV, XLS, or XLSX Files")
uploaded_files = st.file_uploader(
    "Selected files",
    type=["csv", "xls", "xlsx"],
    accept_multiple_files=True
)

start_dates = []

if uploaded_files:
    file_data = {}
    
    # Track which files were uploaded to clean up stale stacks
    uploaded_names = set()
    
    for file in uploaded_files:
        raw_bytes = file.getvalue()
        df_preview, header_row = load_file(raw_bytes, file.name)
        df_preview_short = df_preview.head(2)
        uploaded_names.add(file.name)
        
        if header_row > 0:
            st.caption(f"Detected data starting on line {header_row + 1} in {file.name}")

        file_data[file.name] = {
            "preview": df_preview_short,
            "columns": get_selectable_columns(df_preview),
            "column_types": get_column_types(df_preview, get_selectable_columns(df_preview)),
            "raw_bytes": raw_bytes,
            "selected_cols": {},
            "units": {},
            "cleanup": {},
            "dupe_handling": "Average values"
        }

        # Check for duplicate timestamps
        dt_col = detect_datetime_column(df_preview_short)
        if dt_col:
            try:
                temp_dt = pd.to_datetime(df_preview_short[dt_col], errors="coerce")
                if temp_dt.duplicated().any():
                    with st.expander(f"⚠️ Duplicate timestamps detected in {file.name}"):
                        dupe_choice = st.radio(
                            f"How should duplicates be handled in {file.name}?",
                            ["Average values", "Maximum value", "Minimum value"],
                            key=f"dupes_{file.name}"
                        )
                        file_data[file.name]["dupe_handling"] = dupe_choice
                
                earliest = temp_dt.min()
                if pd.notnull(earliest):
                    start_dates.append(earliest)
            except Exception:
                pass

    # Clean up stacks that reference files no longer uploaded
    stacks_to_remove = []
    for stack_name, stack_info in st.session_state.stacked_datasets.items():
        if not all(fn in uploaded_names for fn in stack_info["files"]):
            stacks_to_remove.append(stack_name)
    
    for stack_name in stacks_to_remove:
        del st.session_state.stacked_datasets[stack_name]
    
    if stacks_to_remove:
        st.warning(f"Removed stacks with missing files: {', '.join(stacks_to_remove)}")
    
    # Restore existing stacked datasets to file_data
    for stack_name, stack_info in st.session_state.stacked_datasets.items():
        # Recreate the stacked dataframe
        stacked_df, _ = create_stacked_dataframe(file_data, stack_info["files"], stack_info["overlap_handling"])
        
        if not stacked_df.empty and len(stacked_df.columns) > 0:
            stacked_df_reset = stacked_df.reset_index()
            if stacked_df_reset.columns[0] == 'index':
                stacked_df_reset = stacked_df_reset.rename(columns={'index': 'DateTime'})
            
            stacked_bytes = stacked_df_reset.to_csv(index=False).encode('utf-8')
            selectable_cols = get_selectable_columns(stacked_df_reset)
            
            # Preserve any previously selected columns/settings if they exist
            old_meta = st.session_state.file_data.get(stack_name, {})
            
            file_data[stack_name] = {
                "preview": stacked_df_reset.head(2),
                "columns": selectable_cols,
                "column_types": get_column_types(stacked_df_reset, selectable_cols),
                "raw_bytes": stacked_bytes,
                "selected_cols": old_meta.get("selected_cols", {}),
                "units": old_meta.get("units", {}),
                "cleanup": old_meta.get("cleanup", {}),
                "dupe_handling": old_meta.get("dupe_handling", "Average values"),
                "is_stacked": True
            }
    
    st.session_state.file_data = file_data
    st.success(f"✅ Uploaded {len(uploaded_files)} files!")

file_data = st.session_state.file_data

# --- Step 1.5: Stack Files from Same Equipment (Optional) ---
if file_data and len(file_data) > 1:
    st.header("📚 Stack Files from Same Equipment (Optional)")
    st.write("If you have multiple files from the same equipment covering different time periods, you can combine them into a single dataset here.")
    
    # Show existing stacks
    if st.session_state.stacked_datasets:
        st.subheader("Current Stacked Datasets")
        stacks_to_remove = []
        
        for stack_name, stack_info in st.session_state.stacked_datasets.items():
            files_in_stack = stack_info["files"]
            col1, col2 = st.columns([4, 1])
            
            with col1:
                # Calculate date range for display
                time_ranges = []
                for fn in files_in_stack:
                    if fn in file_data:
                        df, _ = load_file(file_data[fn]["raw_bytes"], fn)
                        dt_col = detect_datetime_column(df)
                        if dt_col:
                            times = pd.to_datetime(df[dt_col], errors='coerce').dropna()
                            if len(times) > 0:
                                time_ranges.append((times.min(), times.max()))
                
                if time_ranges:
                    overall_start = min(t[0] for t in time_ranges)
                    overall_end = max(t[1] for t in time_ranges)
                    date_str = f"{overall_start.strftime('%Y-%m-%d')} to {overall_end.strftime('%Y-%m-%d')}"
                else:
                    date_str = "Unknown date range"
                
                st.info(f"📊 **{stack_name}**: {', '.join(files_in_stack)} — {date_str}")
            
            with col2:
                if st.button("🗑️ Remove", key=f"remove_stack_{stack_name}"):
                    stacks_to_remove.append(stack_name)
        
        # Remove marked stacks
        for stack_name in stacks_to_remove:
            del st.session_state.stacked_datasets[stack_name]
            st.rerun()
    
    # Create new stack
    with st.expander("➕ Create a new stacked dataset", expanded=not st.session_state.stacked_datasets):
        # Get files not already in a stack
        files_in_stacks = set()
        for stack_info in st.session_state.stacked_datasets.values():
            files_in_stacks.update(stack_info["files"])
        
        available_files = [fn for fn in file_data.keys() if fn not in files_in_stacks]
        
        if len(available_files) < 2:
            st.info("You need at least 2 unstacked files to create a new stack.")
        else:
            stack_files = st.multiselect(
                "Select files to stack together:",
                options=available_files,
                key="new_stack_files"
            )
            
            if len(stack_files) >= 2:
                # Check compatibility
                compat = check_stack_compatibility(file_data, stack_files)
                
                # Show warnings
                for warning in compat["warnings"]:
                    st.warning(warning)
                
                # Show time ranges
                if compat.get("time_ranges"):
                    st.write("**Time ranges in selected files:**")
                    for fn, start, end in compat["time_ranges"]:
                        st.caption(f"  • {fn}: {start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')}")
                
                col1, col2 = st.columns(2)
                with col1:
                    stack_name = st.text_input(
                        "Name for this stacked dataset:",
                        value=f"Stacked Data {len(st.session_state.stacked_datasets) + 1}",
                        key="new_stack_name"
                    )
                
                with col2:
                    overlap_handling = st.selectbox(
                        "If timestamps overlap:",
                        ["Keep first occurrence", "Keep last occurrence", "Average values"],
                        key="new_stack_overlap"
                    )
                
                if st.button("✅ Create Stack", key="create_stack_btn"):
                    if stack_name and stack_name not in st.session_state.stacked_datasets:
                        # Create the stacked dataset
                        stacked_df, _ = create_stacked_dataframe(file_data, stack_files, overlap_handling)
                        
                        if stacked_df.empty or len(stacked_df.columns) == 0:
                            st.error("❌ Could not create stack: No common data columns found between the selected files.")
                        else:
                            # Reset index to make datetime a regular column for storage
                            stacked_df_reset = stacked_df.reset_index()
                            
                            # Rename the index column to something recognizable as datetime
                            if stacked_df_reset.columns[0] == 'index':
                                stacked_df_reset = stacked_df_reset.rename(columns={'index': 'DateTime'})
                            
                            # Convert to CSV bytes for storage
                            stacked_bytes = stacked_df_reset.to_csv(index=False).encode('utf-8')
                            
                            # Get selectable columns (excludes the datetime column)
                            selectable_cols = get_selectable_columns(stacked_df_reset)
                            
                            # Store stack info
                            st.session_state.stacked_datasets[stack_name] = {
                                "files": stack_files,
                                "overlap_handling": overlap_handling
                            }
                            
                            # Add stacked dataset to file_data as a virtual file
                            file_data[stack_name] = {
                                "preview": stacked_df_reset.head(2),
                                "columns": selectable_cols,
                                "column_types": get_column_types(stacked_df_reset, selectable_cols),
                                "raw_bytes": stacked_bytes,
                                "selected_cols": {},
                                "units": {},
                                "cleanup": {},
                                "dupe_handling": "Average values",
                                "is_stacked": True
                            }
                            
                            st.session_state.file_data = file_data
                            st.success(f"✅ Created stacked dataset '{stack_name}' with {len(selectable_cols)} available columns!")
                            st.rerun()
                    elif stack_name in st.session_state.stacked_datasets:
                        st.error("A stack with this name already exists. Please choose a different name.")
                    else:
                        st.error("Please enter a name for the stacked dataset.")
            elif len(stack_files) == 1:
                st.info("Select at least 2 files to stack together.")

file_data = st.session_state.file_data

# --- Step 2: Column Selection & Cleanup ---
if file_data:
    st.header("🧹 Select Columns, Data Titles, & Cleanup Options")
    
    # Get files that are part of stacks (to hide them)
    files_in_stacks = set()
    for stack_info in st.session_state.stacked_datasets.values():
        files_in_stacks.update(stack_info["files"])

    for file_name, meta in file_data.items():
        # Skip files that are part of a stack (they're represented by the stacked dataset)
        if file_name in files_in_stacks:
            continue
            
        df_preview = meta["preview"]
        is_stacked = meta.get("is_stacked", False)
        
        # Different icon for stacked vs regular files
        icon = "📚" if is_stacked else "📈"
        label = f"{icon} {'Stacked: ' if is_stacked else ''}{file_name}"

        with st.expander(f"{label}", expanded=False):
            st.dataframe(df_preview)

            st.markdown(
                "<span style='font-size:20px; font-weight:500;'>"
                "Select data columns to include in the combination</span>",
                unsafe_allow_html=True
            )
            selected = st.multiselect(
                "Select from the list below:",
                options=meta["columns"],
                key=f"select_{file_name}"
            )

            meta["selected_cols"] = {}
            meta["cleanup"] = {}
            meta["units"] = {}

            for col in selected:
                col_type = meta.get("column_types", {}).get(col, "numeric")
                is_string = col_type == "string"
                
                with st.expander(f"⚙️ Settings for {col}" + (" 📝" if is_string else ""), expanded=False):
                    if is_string:
                        st.warning(f"⚠️ '{col}' contains text/string data, not numbers. It will be preserved as text.")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        title = st.text_input(
                            "Data Column Title",
                            value=col,
                            key=f"title_{file_name}_{col}"
                        )
                        meta["selected_cols"][col] = title
                    with col2:
                        units = st.text_input(
                            "Units",
                            value="",
                            key=f"units_{file_name}_{col}"
                        )
                        meta["units"][col] = units
                    with col3:
                        if is_string:
                            # Limited cleanup options for string columns
                            cleanup_choice = st.selectbox(
                                "Missing data handling method",
                                [
                                    "Fill using last valid entry",
                                    "Fill using next available value",
                                    "Delete the entire row of data",
                                    "Fill with empty text"
                                ],
                                key=f"cleanup_{file_name}_{col}"
                            )
                            # Map "Fill with empty text" to "Fill with zero" internally
                            if cleanup_choice == "Fill with empty text":
                                cleanup_choice = "Fill with zero"
                        else:
                            cleanup_choice = st.selectbox(
                                "Missing data handling method",
                                [
                                    "Fill using last valid entry",
                                    "Fill using next available value",
                                    "Fill with a linear interpolation between the nearest values",
                                    "Delete the entire row of data",
                                    "Fill with zero"
                                ],
                                key=f"cleanup_{file_name}_{col}"
                            )
                        meta["cleanup"][col] = cleanup_choice

# --- Step 2.5: Combined Graphing Section (FIXED) ---
if file_data:
    st.header("📊 Visualize Combined Data (Optional)")

    # Build list of available columns from all files (numeric only for graphing)
    available_columns = []
    string_columns_skipped = []
    
    # Get files that are part of stacks (to exclude them)
    files_in_stacks = set()
    for stack_info in st.session_state.stacked_datasets.values():
        files_in_stacks.update(stack_info["files"])
    
    for file_name, meta in file_data.items():
        # Skip files that are part of a stack
        if file_name in files_in_stacks:
            continue
            
        for col, title in meta.get("selected_cols", {}).items():
            col_type = meta.get("column_types", {}).get(col, "numeric")
            if col_type == "string":
                string_columns_skipped.append(title)
                continue
            available_columns.append({
                "file_name": file_name,
                "col": col,
                "title": title,
                "label": title
            })

    if not available_columns:
        if string_columns_skipped:
            st.info(f"No numeric columns available for graphing. String columns ({', '.join(string_columns_skipped)}) cannot be plotted.")
        else:
            st.info("No columns selected yet. Choose some in the previous step.")
    else:
        # Column selection OUTSIDE the button (fixes the Streamlit rerun issue)
        graph_labels = [item["label"] for item in available_columns]
        selected_labels = st.multiselect(
            "Select columns to display on the combined graph:",
            options=graph_labels,
            default=[],
            key="graph_column_select"
        )

        if st.button("Generate Graph") and selected_labels:
            # Group columns by file to load each file only ONCE
            files_to_columns = {}
            for item in available_columns:
                if item["label"] in selected_labels:
                    fn = item["file_name"]
                    if fn not in files_to_columns:
                        files_to_columns[fn] = []
                    files_to_columns[fn].append(item)

            all_series = {}
            progress = st.progress(0, text="Loading data for plotting...")
            
            total_files = len(files_to_columns)
            for i, (file_name, columns) in enumerate(files_to_columns.items(), start=1):
                meta = file_data[file_name]
                
                # Load file ONCE per file (not per column)
                is_stacked = meta.get("is_stacked", False)
                df_full, _ = load_file(meta["raw_bytes"], file_name, is_stacked=is_stacked)
                df_full = prepare_dataframe(df_full, meta.get("dupe_handling", "Average values"))
                
                # Build column mapping (case-insensitive, strip whitespace)
                col_map = {c.lower().strip(): c for c in df_full.columns}
                
                # Extract all requested columns from this file
                for item in columns:
                    col = item["col"]
                    label = item["label"]
                    
                    # Try exact match first, then case-insensitive
                    if col in df_full.columns:
                        matched_col = col
                    else:
                        matched_col = col_map.get(col.lower().strip())
                    
                    if matched_col and matched_col in df_full.columns:
                        series_data = pd.to_numeric(df_full[matched_col], errors="coerce")
                        if not series_data.dropna().empty:
                            # Store as a Series with its datetime index
                            all_series[label] = series_data

                progress.progress(i / total_files, text=f"Loaded {i} of {total_files} files")
            
            # Combine all series into a single DataFrame with aligned index
            if all_series:
                combined_plot_df = pd.DataFrame(all_series)
                combined_plot_df = combined_plot_df.sort_index()
            else:
                combined_plot_df = pd.DataFrame()

            progress.progress(1.0, text="✅ Data ready for plotting")

            if combined_plot_df.empty:
                st.warning("No valid data found for the selected columns.")
            else:
                combined_plot_df = combined_plot_df.sort_index()
                
                # Debug info
                st.caption(f"📊 Plotting {len(combined_plot_df.columns)} columns with {len(combined_plot_df)} data points")
                
                # Melt the dataframe for better plotly express handling
                plot_df = combined_plot_df.reset_index()
                plot_df = plot_df.rename(columns={plot_df.columns[0]: 'DateTime'})
                plot_df_melted = plot_df.melt(id_vars=['DateTime'], var_name='Column', value_name='Value')
                
                # Remove NaN values for cleaner plotting
                plot_df_melted = plot_df_melted.dropna(subset=['Value'])
                
                fig = px.scatter(
                    plot_df_melted,
                    x='DateTime',
                    y='Value',
                    color='Column',
                    title="Combined Data Plot (All Files)",
                )
                fig.update_layout(
                    height=600,
                    hovermode="x unified",
                    legend=dict(orientation="h", y=-0.25)
                )
                st.session_state.graph_fig = fig

    # Display graph if it exists
    if st.session_state.graph_fig:
        st.plotly_chart(st.session_state.graph_fig, use_container_width=True)

# --- Step 3: Time & Interval ---
if file_data:
    st.header("🕐 Time Range & Interval")
    st.write("Select the time range and interval for your combined data set.")
    
    if start_dates:
        default_first = max(start_dates) + pd.Timedelta(days=1)
    else:
        default_first = pd.Timestamp.today().normalize()
    default_last = default_first + pd.Timedelta(days=14)

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("🗓️ Start date", value=default_first)
        range_mode = st.radio(
            "Select end range mode",
            ["End date", "Duration (days)"],
            horizontal=True
        )
    with col2:
        start_time = st.time_input("Start Time", value=datetime.strptime("00:00", "%H:%M").time())
    
    start_dt = datetime.combine(start_date, start_time)

    if range_mode == "End date":
        col1, col2 = st.columns(2)
        with col1:
            end_date = st.date_input("End date", value=default_last)
        with col2:
            end_time = st.time_input("End Time", value=datetime.strptime("00:00", "%H:%M").time())
        end_dt = datetime.combine(end_date, end_time)
    else:
        duration_days = st.number_input("Duration in days", min_value=1, value=14)
        end_dt = start_dt + pd.Timedelta(days=duration_days)

    interval = st.selectbox(
        "⏳ Select output time interval",
        ["1s", "30s", "1min", "5min", "10min", "15min", "1h", "1D"],
        index=2
    )

    st.subheader("📉 Interval Alignment Strategy")
    with st.expander("💡 If your data is on the selected time interval, nothing will be changed. If it is not, it will be resampled according to your selections below."):
        alignment_opts = {}
        
        # Get files that are part of stacks (to exclude them)
        files_in_stacks = set()
        for stack_info in st.session_state.stacked_datasets.values():
            files_in_stacks.update(stack_info["files"])
        
        for file_name in file_data.keys():
            # Skip files that are part of a stack
            if file_name in files_in_stacks:
                continue
                
            is_stacked = file_data[file_name].get("is_stacked", False)
            display_name = f"{'📚 Stacked: ' if is_stacked else ''}{file_name}"
            
            alignment_opts[file_name] = st.selectbox(
                f"Data from '{display_name}':",
                [
                    "Fill with the nearest value",
                    "Do a linear interpolation from the nearest values",
                    "Take an average of the available values within the interval"
                ],
                key=f"align_{file_name}"
            )

# --- Step 4: Create Combined File ---
if file_data:
    st.header("⛷️ Create & Download Combined File")
    
    if st.button("Create combined data file"):
        # Validate template exists
        template_path = Path("Analysis Template.xlsx")
        if not template_path.exists():
            st.error(f"❌ Template file not found: {template_path.absolute()}")
            st.info("Please ensure 'Analysis Template.xlsx' is in the same directory as this script.")
            st.stop()

        dt_index = pd.date_range(start=pd.Timestamp(start_dt), end=pd.Timestamp(end_dt), freq=interval)
        combined = pd.DataFrame(index=dt_index)

        row10 = ["Date"]
        row11 = [""]

        progress = st.progress(0, text="Starting file processing...")
        
        # Get files that are part of stacks (to exclude them)
        files_in_stacks = set()
        for stack_info in st.session_state.stacked_datasets.values():
            files_in_stacks.update(stack_info["files"])
        
        # Filter to only process non-stacked source files
        files_to_process = {fn: meta for fn, meta in file_data.items() if fn not in files_in_stacks}
        total_files = len(files_to_process)

        for i, (file_name, meta) in enumerate(files_to_process.items(), start=1):
            if not meta.get("selected_cols"):
                continue
                
            # Load and prepare file
            is_stacked = meta.get("is_stacked", False)
            df, _ = load_file(meta["raw_bytes"], file_name, is_stacked=is_stacked)
            df = prepare_dataframe(df, meta.get("dupe_handling", "Average values"))

            # Apply cleanup choices
            cleaned_count = 0
            for col, method in meta.get("cleanup", {}).items():
                if col in df.columns:
                    missing_before = df[col].isna().sum()
                    col_type = meta.get("column_types", {}).get(col, "numeric")
                    is_string = col_type == "string"
                    df = apply_cleanup(df, col, method, is_string=is_string)
                    if missing_before > 0:
                        cleaned_count += missing_before

            if cleaned_count > 0:
                st.write(f"✅ {file_name}: {cleaned_count} missing values cleaned up.")

            # Alignment + insert into combined
            for col, title in meta.get("selected_cols", {}).items():
                if col not in df.columns:
                    st.warning(f"Column '{col}' not found in {file_name}")
                    continue
                
                col_type = meta.get("column_types", {}).get(col, "numeric")
                is_string = col_type == "string"
                alignment = alignment_opts.get(file_name, "Fill with the nearest value")
                
                if is_string:
                    # For string columns, use nearest value (forward fill to target index)
                    # First, reindex with ffill to get the most recent value at each target timestamp
                    s = df[col].reindex(df.index.union(dt_index)).ffill().reindex(dt_index)
                else:
                    if alignment == "Fill with the nearest value":
                        s = df[col].reindex(dt_index, method="nearest")
                    elif alignment == "Do a linear interpolation from the nearest values":
                        s = df[col].reindex(dt_index).interpolate(method="time")
                    else:  # Take an average
                        s = df[col].resample(interval).mean().reindex(dt_index)

                combined[title] = s.values
                row10.append(title)
                row11.append(meta.get("units", {}).get(col, ""))

            progress.progress(i / total_files, text=f"Processed {i} of {total_files} files...")

        progress.progress(1.0, text="All files processed ✅")
        
        # Remove rows where ALL data columns are NaN (keeps rows with at least some data)
        if not combined.empty:
            combined = combined.dropna(how='all')

        # Insert into Excel template
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb["Analysis"]

            # Grab formatting from template row
            template_formats = {}
            for col_idx in range(2, 2 + combined.shape[1] + 1):
                template_formats[col_idx] = copy(ws.cell(row=13, column=col_idx)._style)

            # Row 10 - Headers
            for c, val in enumerate(row10, start=2):
                ws.cell(row=10, column=c, value=val)

            # Row 11 - Units
            for c, val in enumerate(row11, start=2):
                ws.cell(row=11, column=c, value=val)

            # Row 12 onward - Data
            for r, (ts, row_vals) in enumerate(combined.iterrows(), start=12):
                # Skip if timestamp is NaT
                if pd.isna(ts):
                    continue
                    
                cell = ws.cell(row=r, column=2, value=ts)
                cell._style = copy(template_formats.get(2, cell._style))

                for c, val in enumerate(row_vals, start=3):
                    # Convert NaN to None for cleaner Excel output
                    if pd.isna(val):
                        val = None
                    cell = ws.cell(row=r, column=c, value=val)
                    if c in template_formats:
                        cell._style = copy(template_formats[c])

            # Apply special formatting to first date cell
            cell = ws.cell(row=12, column=2)
            cell.number_format = "m/d/yy"

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            wb.close()

            st.success("✅ Combined file created!")
            st.download_button(
                label="📥 Download Combined File",
                data=output,
                file_name="Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error creating Excel file: {e}")